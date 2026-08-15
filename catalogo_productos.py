import os
import re
import unicodedata
from pathlib import Path
from threading import RLock

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
_ruta_configurada = os.getenv(
    "CATALOGO_PRODUCTOS_PATH",
    str(BASE_DIR / "datos" / "LISTADO PRODUCTOS.xlsx")
)
CATALOGO_PATH = Path(_ruta_configurada)
if not CATALOGO_PATH.is_absolute():
    CATALOGO_PATH = BASE_DIR / CATALOGO_PATH

LINEAS_PRODUCTO = ("INAPEL", "MARFIL", "TOROFIL")

_MAPA_COLUMNAS = {
    "MARFIL": {
        "referencia": "REFERENCIA SIESA",
        "detalle_presentacion": "EXTENSION MARFIL",
        "producto": "DESCRIPCION INTERNA",
        "plu": "PLU MARFIL",
        "unidad": "UNIDAD DE INVENTARIO"
    },
    "INAPEL": {
        "referencia": "REF",
        "detalle_presentacion": "DETALLE EXT 1",
        "producto": "DESCRIPCION INTERNA",
        "plu": "P L U",
        "unidad": "UNIDAD DE MEDIDA"
    },
    "TOROFIL": {
        "referencia": "REFERENCIA",
        "detalle_presentacion": "PRESENTACION",
        "producto": "",
        "plu": "",
        "unidad": ""
    }
}

_CACHE_LOCK = RLock()
_CACHE_FIRMA = None
_CACHE_PRODUCTOS = ()


def _texto(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _normalizar_encabezado(valor):
    texto = unicodedata.normalize("NFKD", _texto(valor))
    texto = texto.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Z0-9]+", " ", texto.upper()).strip()


def normalizar_linea(linea):
    return _normalizar_encabezado(linea).replace(" ", "_")


def normalizar_referencia(referencia):
    return re.sub(r"\s+", "", _texto(referencia).upper())


def _firma_archivo():
    estado = CATALOGO_PATH.stat()
    return estado.st_mtime_ns, estado.st_size


def _valor_columna(fila, columnas, linea, nombre):
    columna = _normalizar_encabezado(_MAPA_COLUMNAS[linea].get(nombre, ""))
    if not columna:
        return ""
    indice = columnas.get(columna)
    if indice is None or indice >= len(fila):
        return ""
    return _texto(fila[indice])


def _leer_catalogo():
    productos = []
    libro = load_workbook(CATALOGO_PATH, read_only=True, data_only=True)

    try:
        for hoja in libro.worksheets:
            linea = normalizar_linea(hoja.title)
            if linea not in LINEAS_PRODUCTO:
                continue

            encabezados = next(hoja.iter_rows(min_row=1, max_row=1, values_only=True), ())
            columnas = {
                _normalizar_encabezado(valor): indice
                for indice, valor in enumerate(encabezados)
                if _normalizar_encabezado(valor)
            }

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                referencia = _valor_columna(fila, columnas, linea, "referencia")
                if not referencia:
                    continue

                productos.append({
                    "linea": linea,
                    "referencia": referencia,
                    "detalle_presentacion": _valor_columna(fila, columnas, linea, "detalle_presentacion"),
                    "producto": _valor_columna(fila, columnas, linea, "producto"),
                    "plu": _valor_columna(fila, columnas, linea, "plu"),
                    "unidad": _valor_columna(fila, columnas, linea, "unidad")
                })
    finally:
        libro.close()

    return productos


def cargar_catalogo(forzar=False):
    """Carga el maestro una vez y lo actualiza si cambia el archivo."""

    global _CACHE_FIRMA, _CACHE_PRODUCTOS

    with _CACHE_LOCK:
        if not CATALOGO_PATH.exists():
            raise FileNotFoundError(
                f"No existe el catálogo maestro: {CATALOGO_PATH}"
            )

        firma = _firma_archivo()
        if not forzar and _CACHE_FIRMA == firma:
            return list(_CACHE_PRODUCTOS)

        productos = _leer_catalogo()
        _CACHE_FIRMA = firma
        _CACHE_PRODUCTOS = tuple(productos)
        return list(_CACHE_PRODUCTOS)


def buscar_productos(linea, referencia):
    linea_normalizada = normalizar_linea(linea)
    if linea_normalizada not in LINEAS_PRODUCTO:
        raise ValueError("La línea de producto no es válida.")

    referencia_normalizada = normalizar_referencia(referencia)
    if not referencia_normalizada:
        return []

    return [
        dict(producto)
        for producto in cargar_catalogo()
        if producto["linea"] == linea_normalizada
        and normalizar_referencia(producto["referencia"]) == referencia_normalizada
    ]


def recargar_catalogo():
    return cargar_catalogo(forzar=True)
