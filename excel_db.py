import os
import json
import shutil
import hashlib
import tempfile
import threading
import zipfile
from contextlib import contextmanager
from datetime import datetime
from collections import Counter
from functools import wraps
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

print(">>> USANDO excel_db.py <<<")

ARCHIVO = "BaseDatos_PQR.xlsx"

_DB_MUTEX = threading.RLock()
_DB_LOCK_STATE = threading.local()
_DB_LOCK_PATH = os.path.join(
    tempfile.gettempdir(),
    "inapel_pqr_" + hashlib.sha256(os.path.abspath(ARCHIVO).encode("utf-8")).hexdigest()[:16] + ".lock"
)


def _bloquear_archivo(handle):
    if os.name == "nt":
        import msvcrt

        handle.seek(0, os.SEEK_END)
        if handle.tell() == 0:
            handle.write(b"0")
            handle.flush()
        handle.seek(0)
        msvcrt.locking(handle.fileno(), msvcrt.LK_LOCK, 1)
    else:
        import fcntl

        fcntl.flock(handle.fileno(), fcntl.LOCK_EX)


def _desbloquear_archivo(handle):
    if os.name == "nt":
        import msvcrt

        handle.seek(0)
        msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
    else:
        import fcntl

        fcntl.flock(handle.fileno(), fcntl.LOCK_UN)


@contextmanager
def bloqueo_base_datos():
    """Bloquea las escrituras en el proceso y entre procesos del mismo host."""

    with _DB_MUTEX:
        profundidad = getattr(_DB_LOCK_STATE, "profundidad", 0)
        handle = None

        if profundidad == 0:
            try:
                handle = open(_DB_LOCK_PATH, "a+b")
                _bloquear_archivo(handle)
                _DB_LOCK_STATE.handle = handle
            except Exception:
                if handle:
                    handle.close()
                raise

        _DB_LOCK_STATE.profundidad = profundidad + 1

        try:
            yield
        finally:
            _DB_LOCK_STATE.profundidad = profundidad
            if profundidad == 0:
                handle = getattr(_DB_LOCK_STATE, "handle", None)
                try:
                    if handle:
                        _desbloquear_archivo(handle)
                finally:
                    if handle:
                        handle.close()
                    if hasattr(_DB_LOCK_STATE, "handle"):
                        del _DB_LOCK_STATE.handle


def proteger_escritura(func):
    """Ejecuta una operación de escritura bajo el bloqueo compartido."""

    @wraps(func)
    def wrapper(*args, **kwargs):
        with bloqueo_base_datos():
            return func(*args, **kwargs)

    return wrapper


def cargar_workbook_seguro(ruta=None, **opciones):
    """Abre un XLSX existente y deja que BadZipFile detenga la operación."""

    ruta = ruta or ARCHIVO
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No existe la base de datos: {ruta}")

    try:
        return load_workbook(ruta, **opciones)
    except zipfile.BadZipFile:
        print(f"[excel] Base de datos XLSX corrupta, no se sobrescribirá: {ruta}")
        raise


def guardar_workbook_atomico(wb, ruta=None):
    """Guarda, valida y reemplaza un XLSX sin escribir sobre el original."""

    ruta = ruta or ARCHIVO
    with bloqueo_base_datos():
        temporal = None
        directorio = os.path.dirname(os.path.abspath(ruta)) or "."

        try:
            descriptor, temporal = tempfile.mkstemp(
                prefix=f".{os.path.basename(ruta)}.",
                suffix=".xlsx",
                dir=directorio
            )
            os.close(descriptor)

            wb.save(temporal)
            wb.close()

            validacion = cargar_workbook_seguro(temporal, read_only=True, data_only=False)
            validacion.close()

            os.replace(temporal, ruta)
            temporal = None
        finally:
            try:
                wb.close()
            except Exception:
                pass
            if temporal and os.path.exists(temporal):
                try:
                    os.remove(temporal)
                except OSError:
                    pass

HERRAMIENTAS_ANALISIS = (
    "5 ¿Por qué?",
    "Diagrama Ishikawa",
    "Análisis Pareto",
    "Inspección visual",
    "Ensayos de laboratorio",
    "Comparación muestra patrón",
    "Checklist de inspección"
)


def normalizar_herramientas(valor):
    """Convierte el texto legacy o el JSON nuevo en una lista de herramientas."""

    if isinstance(valor, list):
        valores = valor
    else:
        texto = str(valor or "").strip()
        if not texto:
            return []
        try:
            valores = json.loads(texto) if texto.startswith("[") else [texto]
        except (TypeError, ValueError):
            valores = [texto]

    return [str(item).strip() for item in valores if str(item or "").strip()]


def serializar_herramientas(herramientas):
    valores = normalizar_herramientas(herramientas)
    if len(valores) <= 1:
        return valores[0] if valores else ""
    return json.dumps(valores, ensure_ascii=False)


# ==========================================================
# CREAR ARCHIVO Y HOJAS
# ==========================================================

@proteger_escritura
def crear_excel():

    if os.path.exists(ARCHIVO):
        return

    wb = Workbook()

    # ---------------- PQR ----------------

    ws = wb.active
    ws.title = "PQR"

    ws.append([
        "Radicado",
        "Fecha",
        "Hora",
        "Tipo",
        "Cliente",
        "Nit",
        "Contacto",
        "Telefono",
        "Correo",
        "Estado",
        "Prioridad",
        "Descripcion",
        "Expectativa",
        "ProductosJSON",
        "Empresa",
        "Vendedor",
        "Linea",
        "UsuarioID",
        "CorreoConfirmacionEnviado",
        "DocumentoReceptor",
        "CorreoReceptor",
        "TelefonoReceptor",
        "CargoReceptor",
        "AreaReceptor",
        "CiudadRecepcion",
        "DepartamentoRecepcion",
        "MedioRecepcion",
        "OtroMedioRecepcion"
    ])

    # ---------------- Historial ----------------

    ws = wb.create_sheet("Historial")

    ws.append([
        "Radicado",
        "Estado",
        "Usuario",
        "Fecha",
        "Hora",
        "Observacion"
    ])

    # ---------------- Investigaciones ----------------

    ws = wb.create_sheet("Investigaciones")

    ws.append([
        "Radicado",
        "Responsable",
        "Cargo",
        "Herramienta",
        "Causa",
        "Accion",
        "Notificar",
        "FechaRespuesta",
        "FechaCierre",
        "Cierre",
        "Respuesta",
        "Departamentos",
        "CalidadEstado",
        "ComercialEstado",
        "NotificacionComercialEnviada",
        "RespuestaCalidad",
        "RespuestaComercial"
    ])

    # ---------------- Adjuntos ----------------

    ws = wb.create_sheet("Adjuntos")

    ws.append([
        "Radicado",
        "Tipo",
        "Archivo",
        "Ruta",
        "Fecha",
        "Hora",
        "Usuario",
        "Observacion"
    ])

    guardar_workbook_atomico(wb)


# ==========================================================
# VERIFICAR ESTRUCTURA
# ==========================================================

@proteger_escritura
def actualizar_estructura_excel():

    if not os.path.exists(ARCHIVO):
        crear_excel()
        return

    wb = cargar_workbook_seguro()
    cambios = False

    hojas = wb.sheetnames

    if "PQR" not in hojas:
        ws = wb.create_sheet("PQR")
        cambios = True
        ws.append([
            "Radicado",
            "Fecha",
            "Hora",
            "Tipo",
            "Cliente",
            "Nit",
            "Contacto",
            "Telefono",
            "Correo",
            "Estado",
            "Prioridad",
            "Descripcion",
            "Expectativa",
            "ProductosJSON",
            "Empresa",
            "Vendedor",
            "Linea",
            "UsuarioID",
            "CorreoConfirmacionEnviado",
            "DocumentoReceptor",
            "CorreoReceptor",
            "TelefonoReceptor",
            "CargoReceptor",
            "AreaReceptor",
            "CiudadRecepcion",
            "DepartamentoRecepcion",
            "MedioRecepcion",
            "OtroMedioRecepcion"
        ])
    else:
        ws = wb["PQR"]
        if ws.max_column < 14:
            ws.cell(1, 14).value = "ProductosJSON"
            cambios = True
        if ws.max_column < 15:
            ws.cell(1, 15).value = "Empresa"
            cambios = True
        if ws.max_column < 16:
            ws.cell(1, 16).value = "Vendedor"
            cambios = True
        if ws.max_column < 17:
            ws.cell(1, 17).value = "Linea"
            cambios = True
        if ws.max_column < 18:
            ws.cell(1, 18).value = "UsuarioID"
            cambios = True
        if ws.max_column < 19:
            ws.cell(1, 19).value = "CorreoConfirmacionEnviado"
            cambios = True
        if ws.max_column < 20:
            ws.cell(1, 20).value = "DocumentoReceptor"
            cambios = True
        if ws.max_column < 21:
            ws.cell(1, 21).value = "CorreoReceptor"
            cambios = True
        if ws.max_column < 22:
            ws.cell(1, 22).value = "TelefonoReceptor"
            cambios = True
        if ws.max_column < 23:
            ws.cell(1, 23).value = "CargoReceptor"
            cambios = True
        if ws.max_column < 24:
            ws.cell(1, 24).value = "AreaReceptor"
            cambios = True
        if ws.max_column < 25:
            ws.cell(1, 25).value = "CiudadRecepcion"
            cambios = True
        if ws.max_column < 26:
            ws.cell(1, 26).value = "DepartamentoRecepcion"
            cambios = True
        if ws.max_column < 27:
            ws.cell(1, 27).value = "MedioRecepcion"
            cambios = True
        if ws.max_column < 28:
            ws.cell(1, 28).value = "OtroMedioRecepcion"
            cambios = True

    if "Historial" not in hojas:
        ws = wb.create_sheet("Historial")
        cambios = True
        ws.append([
            "Radicado",
            "Estado",
            "Usuario",
            "Fecha",
            "Hora",
            "Observacion"
        ])

    if "Investigaciones" not in hojas:
        ws = wb.create_sheet("Investigaciones")
        cambios = True
        ws.append([
            "Radicado",
            "Responsable",
            "Cargo",
            "Herramienta",
            "Causa",
            "Accion",
            "Notificar",
            "FechaRespuesta",
            "FechaCierre",
            "Cierre",
            "Respuesta",
            "Departamentos",
            "CalidadEstado",
            "ComercialEstado",
            "NotificacionComercialEnviada",
            "RespuestaCalidad",
            "RespuestaComercial"
        ])
    else:
        ws = wb["Investigaciones"]
        if ws.max_column < 13:
            ws.cell(1, 13).value = "CalidadEstado"
            cambios = True
        if ws.max_column < 14:
            ws.cell(1, 14).value = "ComercialEstado"
            cambios = True
        if ws.max_column < 15:
            ws.cell(1, 15).value = "NotificacionComercialEnviada"
            cambios = True
        if ws.max_column < 16:
            ws.cell(1, 16).value = "RespuestaCalidad"
            cambios = True
        if ws.max_column < 17:
            ws.cell(1, 17).value = "RespuestaComercial"
            cambios = True

        # The legacy response belonged to the original commercial field.
        for fila in range(2, ws.max_row + 1):
            if not ws.cell(fila, 17).value and ws.cell(fila, 11).value:
                ws.cell(fila, 17).value = ws.cell(fila, 11).value
                cambios = True

    if "Adjuntos" not in hojas:
        ws = wb.create_sheet("Adjuntos")
        cambios = True
        ws.append([
            "Radicado",
            "Tipo",
            "Archivo",
            "Ruta",
            "Fecha",
            "Hora",
            "Usuario",
            "Observacion"
        ])
    if cambios:
        guardar_workbook_atomico(wb)
    else:
        wb.close()


# ==========================================================
# RADICADO
# ==========================================================

def generar_radicado():

    actualizar_estructura_excel()

    wb = cargar_workbook_seguro()

    ws = wb["PQR"]

    fila = ws.max_row

    if fila <= 1:
        consecutivo = 1

    else:

        ultimo = ws.cell(row=fila, column=1).value

        try:

            consecutivo = int(str(ultimo).split("-")[-1]) + 1

        except:

            consecutivo = fila

    wb.close()

    return f"PQR-{datetime.now().year}-{consecutivo:04d}"


# ==========================================================
# HISTORIAL
# ==========================================================

@proteger_escritura
def guardar_historial(
    radicado,
    estado,
    usuario="Sistema",
    observacion=""
):

    wb = cargar_workbook_seguro()

    ws = wb["Historial"]

    ahora = datetime.now()

    ws.append([
        radicado,
        estado,
        usuario,
        ahora.strftime("%Y-%m-%d"),
        ahora.strftime("%H:%M:%S"),
        observacion
    ])

    guardar_workbook_atomico(wb)


# ==========================================================
# GUARDAR PQR
# ==========================================================

@proteger_escritura
def guardar_pqr(datos):

    actualizar_estructura_excel()

    wb = cargar_workbook_seguro()

    ws = wb["PQR"]

    productos_json = json.dumps(datos.get("productos", []), ensure_ascii=False)

    ws.append([

        datos.get("radicado",""),
        datos.get("fechaRec",""),
        datos.get("horaRec",""),
        datos.get("tipoSol",""),
        datos.get("cliente",""),
        datos.get("nit",""),
        datos.get("contacto",""),
        datos.get("tel",""),
        datos.get("email",""),
        datos.get("estado","Recibido"),
        datos.get("prioridad",""),
        datos.get("desc",""),
        datos.get("expectativa",""),
        productos_json,
        datos.get("empresa","INAPEL"),
        datos.get("vendedor",""),
        datos.get("linea",""),
        datos.get("usuario_id",""),
        "NO",
        datos.get("documento_receptor", ""),
        datos.get("correo_receptor", ""),
        datos.get("telefono_receptor", ""),
        datos.get("cargo_receptor", ""),
        datos.get("area_receptor", ""),
        datos.get("ciudadRec", ""),
        datos.get("dptoRec", ""),
        datos.get("medio", ""),
        datos.get("otroMedio", "")

    ])

    guardar_workbook_atomico(wb)

    guardar_historial(

        datos["radicado"],
        datos.get("estado","Recibido"),
        "Sistema",
        "PQR registrado"

    )

    return True


# ==========================================================
# CORREO DE CONFIRMACIÓN
# ==========================================================

def correo_confirmacion_enviado(radicado):
    """Devuelve True si el correo de confirmación ya fue enviado para el radicado."""

    wb = cargar_workbook_seguro()
    ws = wb["PQR"]

    for fila in ws.iter_rows(min_row=2):
        if fila[0].value is not None and str(fila[0].value).strip() == str(radicado).strip():
            valor = fila[18].value if len(fila) > 18 else None
            wb.close()
            return str(valor or "").strip().upper() == "SI"

    wb.close()
    return False


@proteger_escritura
def marcar_correo_confirmacion(radicado, enviado):
    """Actualiza la columna CorreoConfirmacionEnviado (SI/NO) del radicado."""

    wb = cargar_workbook_seguro()
    ws = wb["PQR"]

    for fila in ws.iter_rows(min_row=2):
        if fila[0].value is not None and str(fila[0].value).strip() == str(radicado).strip():
            fila[18].value = "SI" if enviado else "NO"
            break

    guardar_workbook_atomico(wb)

# ==========================================================
# CONSULTAR PQR
# ==========================================================

def consultar_pqr(valor_busqueda):

    actualizar_estructura_excel()
    wb = cargar_workbook_seguro()

    ws = wb["PQR"]
    ws_inv = wb["Investigaciones"]
    ws_hist = wb["Historial"]

    valor_busqueda = str(valor_busqueda).strip().upper()

    for fila in ws.iter_rows(min_row=2, values_only=True):

        radicado = str(fila[0]).strip().upper()
        cliente = str(fila[4]).strip().upper()
        nit = str(fila[5]).strip().upper()

        if valor_busqueda in (radicado, cliente, nit):

            investigacion = {
                "resp": "",
                "cargo": "",
                "herr": "5 ¿Por qué?",
                "herramientas": ["5 ¿Por qué?"],
                "causa": "Materias primas",
                "acc": "Reposición",
                "notif": "Sí",
                "fResp": "",
                "fCierre": "",
                "cierre": "No",
                "respuesta_calidad": "",
                "respuesta_comercial": "",
                "deptos": "",
                "calidad_estado": "pendiente",
                "comercial_estado": "pendiente",
                "notificacion_comercial_enviada": False
            }

            for inv in ws_inv.iter_rows(min_row=2, values_only=True):
                if str(inv[0]).strip().upper() == radicado:
                    herramientas = normalizar_herramientas(inv[3])
                    investigacion = {
                        "resp": inv[1] or "",
                        "cargo": inv[2] or "",
                        "herr": herramientas[0] if herramientas else "",
                        "herramientas": herramientas,
                        "causa": inv[4] or "",
                        "acc": inv[5] or "",
                        "notif": inv[6] or "",
                        "fResp": inv[7] or "",
                        "fCierre": inv[8] or "",
                        "cierre": inv[9] or "",
                        "respuesta_calidad": inv[15] or "" if len(inv) > 15 else "",
                        "respuesta_comercial": (inv[16] if len(inv) > 16 and inv[16] else (inv[10] or "")),
                        "deptos": inv[11] or "" if len(inv) > 11 else "",
                        "calidad_estado": str(inv[12] or "pendiente").strip().lower() if len(inv) > 12 else "pendiente",
                        "comercial_estado": str(inv[13] or "pendiente").strip().lower() if len(inv) > 13 else "pendiente",
                        "notificacion_comercial_enviada": str(inv[14] or "").strip().lower() in ("1", "true", "sí", "si", "enviada") if len(inv) > 14 else False
                    }
                    break

            historial = []
            for h in ws_hist.iter_rows(min_row=2, values_only=True):
                if str(h[0]).strip().upper() == radicado:
                    historial.append({
                        "estado": h[1],
                        "fecha": str(h[3]),
                        "hora": str(h[4]),
                        "usuario": str(h[2])
                    })

            productos_raw = fila[13] if len(fila) > 13 and fila[13] else "[]"
            try:
                productos = json.loads(productos_raw)
            except:
                productos = []

            wb.close()
            return {
                "radicado": fila[0],
                "fechaRec": str(fila[1]),
                "horaRec": fila[2],
                "tipoSol": fila[3],
                "cliente": fila[4],
                "nit": fila[5],
                "contacto": fila[6],
                "tel": fila[7],
                "email": fila[8],
                "estado": fila[9],
                "prioridad": fila[10],
                "desc": fila[11],
                "expectativa": fila[12] if len(fila) > 12 else "",
                "productos": productos,
                "empresa": fila[14] if len(fila) > 14 and fila[14] else "",
                "vendedor": fila[15] if len(fila) > 15 and fila[15] else "",
                "linea": fila[16] if len(fila) > 16 and fila[16] else "",
                "usuario_id": fila[17] if len(fila) > 17 and fila[17] else "",
                "documento_receptor": fila[19] if len(fila) > 19 and fila[19] else "",
                "correo_receptor": fila[20] if len(fila) > 20 and fila[20] else "",
                "telefono_receptor": fila[21] if len(fila) > 21 and fila[21] else "",
                "cargo_receptor": fila[22] if len(fila) > 22 and fila[22] else "",
                "area_receptor": fila[23] if len(fila) > 23 and fila[23] else "",
                "ciudad_recepcion": fila[24] if len(fila) > 24 and fila[24] else "",
                "departamento_recepcion": fila[25] if len(fila) > 25 and fila[25] else "",
                "medio_recepcion": fila[26] if len(fila) > 26 and fila[26] else "",
                "otro_medio_recepcion": fila[27] if len(fila) > 27 and fila[27] else "",
                "investigacion": investigacion,
                "historial": historial,
                "savedAt": f"{fila[1]}T{fila[2]}"
            }

    wb.close()
    return None

# ==========================================================
# LISTAR TODOS LOS PQR (misma fuente que obtener_dashboard)
# ==========================================================

def listar_pqrs():

    actualizar_estructura_excel()
    wb = cargar_workbook_seguro()

    ws = wb["PQR"]

    lista = []

    for fila in ws.iter_rows(min_row=2, values_only=True):

        if not fila[0]:
            continue

        fecha = str(fila[1])[:10] if fila[1] else ""
        hora = fila[2] or ""

        productos_raw = fila[13] if len(fila) > 13 and fila[13] else "[]"

        try:
            productos = json.loads(productos_raw)
        except:
            productos = []

        lista.append({
            "radicado": fila[0],
            "fechaRec": fecha,
            "horaRec": hora,
            "tipoSol": fila[3] or "",
            "cliente": fila[4] or "",
            "nit": fila[5] or "",
            "contacto": fila[6] or "",
            "tel": fila[7] or "",
            "email": fila[8] or "",
            "estado": fila[9] or "Recibido",
            "prioridad": fila[10] or "",
            "desc": fila[11] or "",
            "expectativa": fila[12] if len(fila) > 12 else "",
            "productos": productos,
            "empresa": fila[14] if len(fila) > 14 and fila[14] else "",
            "vendedor": fila[15] if len(fila) > 15 and fila[15] else "",
            "linea": fila[16] if len(fila) > 16 and fila[16] else "",
            "usuario_id": fila[17] if len(fila) > 17 and fila[17] else "",
            "documento_receptor": fila[19] if len(fila) > 19 and fila[19] else "",
            "correo_receptor": fila[20] if len(fila) > 20 and fila[20] else "",
            "telefono_receptor": fila[21] if len(fila) > 21 and fila[21] else "",
            "cargo_receptor": fila[22] if len(fila) > 22 and fila[22] else "",
            "area_receptor": fila[23] if len(fila) > 23 and fila[23] else "",
            "ciudad_recepcion": fila[24] if len(fila) > 24 and fila[24] else "",
            "departamento_recepcion": fila[25] if len(fila) > 25 and fila[25] else "",
            "medio_recepcion": fila[26] if len(fila) > 26 and fila[26] else "",
            "otro_medio_recepcion": fila[27] if len(fila) > 27 and fila[27] else "",
            "savedAt": f"{fecha}T{hora}" if hora else fecha
        })

    wb.close()
    return lista

# ==========================================================
# ACTUALIZAR ESTADO DEL PQR
# ==========================================================

@proteger_escritura
def actualizar_estado_pqr(radicado, estado):

    actualizar_estructura_excel()

    wb = cargar_workbook_seguro()

    ws = wb["PQR"]

    for fila in ws.iter_rows(min_row=2):

        if str(fila[0].value).strip() == radicado:

            fila[9].value = estado
            break

    guardar_workbook_atomico(wb)

    return True


# ==========================================================
# GUARDAR / ACTUALIZAR INVESTIGACIÓN
# ==========================================================

@proteger_escritura
def guardar_investigacion(
    datos,
    calidad_estado=None,
    comercial_estado=None,
    notificacion_comercial_enviada=None
):

    actualizar_estructura_excel()

    wb = cargar_workbook_seguro()
    ws = wb["Investigaciones"]
    fila_existente = None

    for fila in range(2, ws.max_row + 1):
        if str(ws.cell(fila, 1).value).strip() == str(datos["radicado"]).strip():
            fila_existente = fila
            break

    def _bool(valor):
        return str(valor or "").strip().lower() in ("1", "true", "sí", "si", "enviada")

    estado_calidad_anterior = (
        ws.cell(fila_existente, 13).value
        if fila_existente and ws.max_column >= 13
        else "pendiente"
    )
    estado_comercial_anterior = (
        ws.cell(fila_existente, 14).value
        if fila_existente and ws.max_column >= 14
        else "pendiente"
    )
    aviso_anterior = (
        ws.cell(fila_existente, 15).value
        if fila_existente and ws.max_column >= 15
        else False
    )
    respuesta_legacy = (
        ws.cell(fila_existente, 11).value
        if fila_existente and ws.max_column >= 11
        else ""
    )

    estado_calidad = str(
        calidad_estado if calidad_estado is not None else estado_calidad_anterior or "pendiente"
    ).strip().lower()
    estado_comercial = str(
        comercial_estado if comercial_estado is not None else estado_comercial_anterior or "pendiente"
    ).strip().lower()
    aviso_enviado = (
        _bool(notificacion_comercial_enviada)
        if notificacion_comercial_enviada is not None
        else _bool(aviso_anterior)
    )
    herramientas = normalizar_herramientas(
        datos.get("herramientas", datos.get("herr", ""))
    )

    valores = [
        datos.get("radicado", ""),
        datos.get("resp", ""),
        datos.get("cargo", ""),
        serializar_herramientas(herramientas),
        datos.get("causa", ""),
        datos.get("acc", ""),
        datos.get("notif", ""),
        datos.get("fResp", ""),
        datos.get("fCierre", ""),
        datos.get("cierre", ""),
        respuesta_legacy,
        datos.get("deptos", "")
    ]
    respuesta_calidad = datos.get("respuesta_calidad", "")
    respuesta_comercial = datos.get("respuesta_comercial", "")

    if fila_existente:
        for columna, valor in enumerate(valores, start=1):
            ws.cell(fila_existente, columna).value = valor
        ws.cell(fila_existente, 13).value = estado_calidad
        ws.cell(fila_existente, 14).value = estado_comercial
        ws.cell(fila_existente, 15).value = 1 if aviso_enviado else 0
        ws.cell(fila_existente, 16).value = respuesta_calidad
        ws.cell(fila_existente, 17).value = respuesta_comercial
    else:
        ws.append(valores + [
            estado_calidad,
            estado_comercial,
            1 if aviso_enviado else 0,
            respuesta_calidad,
            respuesta_comercial
        ])

    guardar_workbook_atomico(wb)

    estado = "Cerrado" if datos.get("cierre") == "Sí" else "En investigación"
    actualizar_estado_pqr(datos["radicado"], estado)
    guardar_historial(
        datos["radicado"],
        estado,
        "Sistema",
        "Seguimiento actualizado"
    )

    return {
        "calidad_estado": estado_calidad,
        "comercial_estado": estado_comercial,
        "notificacion_comercial_enviada": aviso_enviado
    }


@proteger_escritura
def marcar_notificacion_comercial_enviada(radicado):
    """Marca el aviso comercial como enviado sin alterar los demás datos."""

    actualizar_estructura_excel()
    wb = cargar_workbook_seguro()
    ws = wb["Investigaciones"]

    for fila in range(2, ws.max_row + 1):
        if str(ws.cell(fila, 1).value).strip() == str(radicado).strip():
            ws.cell(fila, 15).value = 1
            guardar_workbook_atomico(wb)
            return True

    wb.close()
    return False

# ==========================================================
# INICIALIZAR ESTRUCTURA
# ==========================================================

@proteger_escritura
def inicializar_excel():
    """
    Crea el archivo y las hojas necesarias si no existen.
    Puede llamarse al iniciar la aplicación.
    """
    actualizar_estructura_excel()

    wb = cargar_workbook_seguro()
    cambios = False

    hojas_requeridas = {
        "PQR": [
            "Radicado", "Fecha", "Hora", "Tipo", "Cliente", "Nit",
            "Contacto", "Telefono", "Correo", "Estado",
            "Prioridad", "Descripcion", "Expectativa", "ProductosJSON",
            "Empresa", "Vendedor", "Linea", "UsuarioID",
            "CorreoConfirmacionEnviado", "DocumentoReceptor",
            "CorreoReceptor", "TelefonoReceptor", "CargoReceptor",
            "AreaReceptor", "CiudadRecepcion", "DepartamentoRecepcion",
            "MedioRecepcion", "OtroMedioRecepcion"
        ],
        "Historial": [
            "Radicado", "Estado", "Usuario",
            "Fecha", "Hora", "Observacion"
        ],
        "Investigaciones": [
            "Radicado", "Responsable", "Cargo",
            "Herramienta", "Causa", "Accion",
            "Notificar", "FechaRespuesta",
            "FechaCierre", "Cierre", "Respuesta",
            "Departamentos"
        ],
        "Adjuntos": [
            "Radicado", "Tipo", "Archivo",
            "Ruta", "Fecha", "Hora",
            "Usuario", "Observacion"
        ]
    }

    for hoja, encabezados in hojas_requeridas.items():

        if hoja not in wb.sheetnames:
            ws = wb.create_sheet(hoja)
            ws.append(encabezados)
            cambios = True

    if cambios:
        guardar_workbook_atomico(wb)
    else:
        wb.close()


def obtener_dashboard():

    actualizar_estructura_excel()

    wb = cargar_workbook_seguro()

    ws = wb["PQR"]

    estados = Counter()
    tipos = Counter()
    prioridades = Counter()

    total = 0

    for fila in ws.iter_rows(min_row=2, values_only=True):

        total += 1

        estado = fila[9] or "Sin estado"
        tipo = fila[3] or "Sin tipo"
        prioridad = fila[10] or "Sin prioridad"

        estados[estado] += 1
        tipos[tipo] += 1
        prioridades[prioridad] += 1

    wb.close()

    return {
        "total": total,
        "estados": dict(estados),
        "tipos": dict(tipos),
        "prioridades": dict(prioridades)
    }
# ==========================================================
# GUARDAR ADJUNTO
# ==========================================================

@proteger_escritura
def guardar_adjunto(
    radicado,
    tipo,
    archivo_original,
    ruta_archivo,
    observacion="",
    usuario="Cliente"
):

    actualizar_estructura_excel()

    wb = cargar_workbook_seguro()

    ws = wb["Adjuntos"]

    ahora = datetime.now()

    fila = ws.max_row + 1

    # Radicado
    ws.cell(fila, 1).value = radicado

    # Tipo
    ws.cell(fila, 2).value = tipo

    # Archivo original
    ws.cell(fila, 3).value = archivo_original

    # Hipervínculo
    celda = ws.cell(fila, 4)
    celda.value = "📎 Abrir evidencia"
    celda.hyperlink = ruta_archivo
    celda.font = Font(underline="single", color="0000FF")

    # Fecha
    ws.cell(fila, 5).value = ahora.strftime("%Y-%m-%d")

    # Hora
    ws.cell(fila, 6).value = ahora.strftime("%H:%M:%S")

    # Usuario
    ws.cell(fila, 7).value = usuario

    # Observacion
    ws.cell(fila, 8).value = observacion

    guardar_workbook_atomico(wb)

    return True

# ==========================================================
# ELIMINAR PQR
# ==========================================================

@proteger_escritura
def eliminar_pqr(radicado):

    actualizar_estructura_excel()

    wb = cargar_workbook_seguro()

    ws = wb["PQR"]

    encontrado = False

    for fila in range(2, ws.max_row + 1):

        if str(ws.cell(fila, 1).value).strip().upper() == str(radicado).strip().upper():

            ws.delete_rows(fila)
            encontrado = True
            break

    if not encontrado:

        wb.close()
        return "not_found"

    for hoja in ["Historial", "Investigaciones", "Adjuntos"]:

        if hoja not in wb.sheetnames:
            continue

        ws_h = wb[hoja]

        filas_a_eliminar = []

        for fila in range(2, ws_h.max_row + 1):

            if str(ws_h.cell(fila, 1).value).strip().upper() == str(radicado).strip().upper():
                filas_a_eliminar.append(fila)

        for fila in reversed(filas_a_eliminar):
            ws_h.delete_rows(fila)

    try:

        guardar_workbook_atomico(wb)

    except Exception:

        return "save_error"

    carpeta = os.path.join("Base_Datos", "Evidencias", str(radicado))

    if os.path.isdir(carpeta):
        shutil.rmtree(carpeta, ignore_errors=True)

    return True
