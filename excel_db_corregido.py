import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

ARCHIVO = "BaseDatos_PQR.xlsx"


# ==========================================================
# CREAR ARCHIVO Y HOJAS
# ==========================================================

def crear_excel():

    if os.path.exists(ARCHIVO):
        return

    wb = Workbook()

    # ---------------- PQR ----------------

    ws = wb.active
    ws.title = "PQR"

    ws.append([
        "Radicado",
        "Fecha",
        "Hora",
        "Tipo",
        "Cliente",
        "Nit",
        "Contacto",
        "Telefono",
        "Correo",
        "Estado",
        "Prioridad",
        "Descripcion",
        "Expectativa"
    ])

    # ---------------- Historial ----------------

    ws = wb.create_sheet("Historial")

    ws.append([
        "Radicado",
        "Estado",
        "Usuario",
        "Fecha",
        "Hora",
        "Observacion"
    ])

    # ---------------- Investigaciones ----------------

    ws = wb.create_sheet("Investigaciones")

    ws.append([
        "Radicado",
        "Responsable",
        "Cargo",
        "Herramienta",
        "Causa",
        "Accion",
        "Notificar",
        "FechaRespuesta",
        "FechaCierre",
        "Cierre",
        "Respuesta"
    ])

    # ---------------- Adjuntos ----------------

    ws = wb.create_sheet("Adjuntos")

    ws.append([
        "Radicado",
        "Archivo",
        "Ruta",
        "Fecha",
        "Hora",
        "Usuario"
    ])

    wb.save(ARCHIVO)
    wb.close()


# ==========================================================
# VERIFICAR ESTRUCTURA
# ==========================================================

def actualizar_estructura_excel():

    if not os.path.exists(ARCHIVO):
        crear_excel()


# ==========================================================
# RADICADO
# ==========================================================

def generar_radicado():

    actualizar_estructura_excel()

    wb = load_workbook(ARCHIVO)

    ws = wb["PQR"]

    fila = ws.max_row

    if fila <= 1:
        consecutivo = 1

    else:

        ultimo = ws.cell(row=fila, column=1).value

        try:

            consecutivo = int(str(ultimo).split("-")[-1]) + 1

        except:

            consecutivo = fila

    wb.close()

    return f"PQR-{datetime.now().year}-{consecutivo:04d}"


# ==========================================================
# HISTORIAL
# ==========================================================

def guardar_historial(
    radicado,
    estado,
    usuario="Sistema",
    observacion=""
):

    wb = load_workbook(ARCHIVO)

    ws = wb["Historial"]

    ahora = datetime.now()

    ws.append([
        radicado,
        estado,
        usuario,
        ahora.strftime("%Y-%m-%d"),
        ahora.strftime("%H:%M:%S"),
        observacion
    ])

    wb.save(ARCHIVO)
    wb.close()


# ==========================================================
# GUARDAR PQR
# ==========================================================

def guardar_pqr(datos):

    actualizar_estructura_excel()

    wb = load_workbook(ARCHIVO)

    ws = wb["PQR"]

    ws.append([

        datos.get("radicado",""),
        datos.get("fechaRec",""),
        datos.get("horaRec",""),
        datos.get("tipoSol",""),
        datos.get("cliente",""),
        datos.get("nit",""),
        datos.get("contacto",""),
        datos.get("tel",""),
        datos.get("email",""),
        datos.get("estado","Recibido"),
        datos.get("prioridad",""),
        datos.get("desc",""),
        datos.get("expectativa","")

    ])

    wb.save(ARCHIVO)
    wb.close()

    guardar_historial(

        datos["radicado"],
        datos.get("estado","Recibido"),
        "Sistema",
        "PQR registrado"

    )

    return True

# ==========================================================
# CONSULTAR PQR
# ==========================================================

def consultar_pqr(valor_busqueda):

    actualizar_estructura_excel()
    wb = load_workbook(ARCHIVO)

    ws = wb["PQR"]
    ws_inv = wb["Investigaciones"]
    ws_hist = wb["Historial"]

    valor_busqueda = str(valor_busqueda).strip().upper()

    for fila in ws.iter_rows(min_row=2, values_only=True):

        radicado = str(fila[0]).strip().upper()
        cliente = str(fila[4]).strip().upper()
        nit = str(fila[5]).strip().upper()

        if valor_busqueda in (radicado, cliente, nit):

            investigacion = {}

            for inv in ws_inv.iter_rows(min_row=2, values_only=True):
                if str(inv[0]).strip().upper() == radicado:
                    investigacion = {
                        "resp": inv[1] or "",
                        "cargo": inv[2] or "",
                        "herr": inv[3] or "",
                        "causa": inv[4] or "",
                        "acc": inv[5] or "",
                        "notif": inv[6] or "",
                        "fResp": inv[7] or "",
                        "fCierre": inv[8] or "",
                        "cierre": inv[9] or "",
                        "respTxt": inv[10] or ""
                    }
                    break

            historial = []
            for h in ws_hist.iter_rows(min_row=2, values_only=True):
                if str(h[0]).strip().upper() == radicado:
                    historial.append({
                        "estado": h[1],
                        "fecha": str(h[3]),
                        "hora": str(h[4]),
                        "usuario": str(h[2])
                    })

            wb.close()
            return {
                "radicado": fila[0],
                "fechaRec": str(fila[1]),
                "horaRec": fila[2],
                "tipoSol": fila[3],
                "cliente": fila[4],
                "nit": fila[5],
                "contacto": fila[6],
                "tel": fila[7],
                "email": fila[8],
                "estado": fila[9],
                "prioridad": fila[10],
                "desc": fila[11],
                "expectativa": fila[12],
                "investigacion": investigacion,
                "historial": historial,
                "savedAt": f"{fila[1]}T{fila[2]}"
            }

    wb.close()
    return None

# ==========================================================
# ACTUALIZAR ESTADO DEL PQR
# ==========================================================

def actualizar_estado_pqr(radicado, estado):

    actualizar_estructura_excel()

    wb = load_workbook(ARCHIVO)
    import os

        print("=" * 60)
        print("ARCHIVO:", ARCHIVO)
        print("EXISTE :", os.path.exists(ARCHIVO))
        print("HOJAS  :", wb.sheetnames)
        print("=" * 60)

    ws = wb["PQR"]

    for fila in ws.iter_rows(min_row=2):

        if str(fila[0].value).strip() == radicado:

            fila[9].value = estado
            break

    wb.save(ARCHIVO)
    wb.close()

    return True


# ==========================================================
# GUARDAR / ACTUALIZAR INVESTIGACIÓN
# ==========================================================

def guardar_investigacion(datos):

    actualizar_estructura_excel()

    wb = load_workbook(ARCHIVO)

    ws = wb["Investigaciones"]

    fila_existente = None

    # Buscar si ya existe el radicado

    for fila in range(2, ws.max_row + 1):

        if str(ws.cell(fila, 1).value).strip() == str(datos["radicado"]).strip():

            fila_existente = fila
            break

    # Si existe, actualiza

    if fila_existente:

        ws.cell(fila_existente,1).value = datos.get("radicado","")
        ws.cell(fila_existente,2).value = datos.get("resp","")
        ws.cell(fila_existente,3).value = datos.get("cargo","")
        ws.cell(fila_existente,4).value = datos.get("herr","")
        ws.cell(fila_existente,5).value = datos.get("causa","")
        ws.cell(fila_existente,6).value = datos.get("acc","")
        ws.cell(fila_existente,7).value = datos.get("notif","")
        ws.cell(fila_existente,8).value = datos.get("fResp","")
        ws.cell(fila_existente,9).value = datos.get("fCierre","")
        ws.cell(fila_existente,10).value = datos.get("cierre","")
        ws.cell(fila_existente,11).value = datos.get("respTxt","")

    else:

        ws.append([

            datos.get("radicado",""),
            datos.get("resp",""),
            datos.get("cargo",""),
            datos.get("herr",""),
            datos.get("causa",""),
            datos.get("acc",""),
            datos.get("notif",""),
            datos.get("fResp",""),
            datos.get("fCierre",""),
            datos.get("cierre",""),
            datos.get("respTxt","")

        ])

    wb.save(ARCHIVO)
    wb.close()

    # ==============================
    # Estado automático
    # ==============================

    estado = "En investigación"

    if datos.get("cierre") == "Sí":

        estado = "Cerrado"

    actualizar_estado_pqr(

        datos["radicado"],
        estado

    )

    guardar_historial(

        datos["radicado"],
        estado,
        "Sistema",
        "Seguimiento actualizado"

    )

    return True

# ==========================================================
# INICIALIZAR ESTRUCTURA
# ==========================================================

def inicializar_excel():
    """
    Crea el archivo y las hojas necesarias si no existen.
    Puede llamarse al iniciar la aplicación.
    """
    actualizar_estructura_excel()

    wb = load_workbook(ARCHIVO)

    hojas_requeridas = {
        "PQR": [
            "Radicado", "Fecha", "Hora", "Tipo", "Cliente", "Nit",
            "Contacto", "Telefono", "Correo", "Estado",
            "Prioridad", "Descripcion", "Expectativa"
        ],
        "Historial": [
            "Radicado", "Estado", "Usuario",
            "Fecha", "Hora", "Observacion"
        ],
        "Investigaciones": [
            "Radicado", "Responsable", "Cargo",
            "Herramienta", "Causa", "Accion",
            "Notificar", "FechaRespuesta",
            "FechaCierre", "Cierre", "Respuesta"
        ],
        "Adjuntos": [
            "Radicado", "Archivo", "Ruta",
            "Fecha", "Hora", "Usuario"
        ]
    }

    for hoja, encabezados in hojas_requeridas.items():

        if hoja not in wb.sheetnames:
            ws = wb.create_sheet(hoja)
            ws.append(encabezados)

    wb.save(ARCHIVO)
    wb.close()
from collections import Counter
from openpyxl import load_workbook

def obtener_dashboard():

    wb = load_workbook(ARCHIVO)

    ws = wb["PQR"]

    total = 0

    estados = Counter()
    prioridades = Counter()
    tipos = Counter()

    for fila in ws.iter_rows(min_row=2, values_only=True):

        if not fila[0]:
            continue

        total += 1

        estado = str(fila[9] or "")
        prioridad = str(fila[10] or "")
        tipo = str(fila[3] or "")

        estados[estado] += 1
        prioridades[prioridad] += 1
        tipos[tipo] += 1

    wb.close()

    return {
        "total": total,
        "estados": dict(estados),
        "prioridades": dict(prioridades),
        "tipos": dict(tipos)
    }
# ==========================================================
# DASHBOARD
# ==========================================================

from collections import Counter

def obtener_dashboard():

    actualizar_estructura_excel()

    wb = load_workbook(ARCHIVO)

    ws = wb["PQR"]

    estados = Counter()
    tipos = Counter()
    prioridades = Counter()

    total = 0

    for fila in ws.iter_rows(min_row=2, values_only=True):

        total += 1

        estado = fila[9] or "Sin estado"
        tipo = fila[3] or "Sin tipo"
        prioridad = fila[10] or "Sin prioridad"

        estados[estado] += 1
        tipos[tipo] += 1
        prioridades[prioridad] += 1

    wb.close()

    return {
        "total": total,
        "estados": dict(estados),
        "tipos": dict(tipos),
        "prioridades": dict(prioridades)
    }