"""Storage adapter for the PQR XLSX master using Google Apps Script Web App.

The adapter communicates with a Google Apps Script Web App that handles
reading and writing the XLSX file to Google Drive. This allows Render Free
to persist data without PostgreSQL, Persistent Disk, or Google Cloud
Service Accounts.

The real Google credentials (refresh token, client ID/secret) are NOT stored
in this module or in the Flask application. Instead, the Apps Script Web App
handles the Google Drive authentication on the server side.

The Python module only needs:
- PQR_STORAGE_URL: URL base de la Apps Script Web App
- PQR_STORAGE_TOKEN: Token secreto para autorizar las llamadas API

Usage:
    PQR_STORAGE=google_drive
    PQR_STORAGE_URL=https://script.google.com/macros/s/EXEC_ID/exec
    PQR_STORAGE_TOKEN=mi-token-secreto

The adapter uses POST /exec with JSON body containing "token" and "action".
See the API specification below.

ENDPOINTS DE LA API DE APPS SCRIPT:

POST /exec
Body JSON obligatorio:
{
  "token": "<secreto>",
  "action": "get_master" | "save_master" | "backup_master"
}

Respuestas esperadas:

get_master:
{
  "ok": true,
  "action": "get_master",
  "fileName": "BaseDatos_PQR.xlsx",
  "fileId": "...",
  "mimeType": "...",
  "size": 14692,
  "base64": "..."
}

save_master:
{
  "ok": true,
  "action": "save_master"
}

backup_master:
{
  "ok": true,
  "action": "backup_master"
}

The Python module ONLY needs:
- PQR_STORAGE_URL: URL base de la Apps Script (https://script.google.com/macros/s/EXEC_ID/exec)
- PQR_STORAGE_TOKEN: Token secreto compartido entre Flask y la Apps Script
- No Google credentials, no OAuth, no Service Account in the Python code
"""

import hashlib
import json
import logging
import os
import tempfile
from datetime import datetime, timezone
from io import BytesIO

import requests
import zipfile

LOGGER = logging.getLogger(__name__)

XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
REQUIRED_SHEETS = ("PQR", "Historial", "Investigaciones", "Adjuntos", "Usuarios")
PQR_HEADERS = (
    "Radicado",
    "Fecha",
    "Hora",
    "Tipo",
    "Cliente",
    "Nit",
    "Contacto",
    "Telefono",
    "Correo",
    "Estado",
    "Prioridad",
    "Descripcion",
    "Expectativa",
    "ProductosJSON",
    "Empresa",
    "Vendedor",
    "Linea",
    "UsuarioID",
    "CorreoConfirmacionEnviado",
    "DocumentoReceptor",
    "CorreoReceptor",
    "TelefonoReceptor",
    "CargoReceptor",
    "AreaReceptor",
    "CiudadRecepcion",
    "DepartamentoRecepcion",
    "MedioRecepcion",
    "OtroMedioRecepcion"
)


class StorageError(RuntimeError):
    """Base error for storage failures."""


class StorageConfigurationError(StorageError):
    """Raised when storage configuration is incomplete."""


class StorageUnavailableError(StorageError):
    """Raised when the storage service cannot complete an operation."""


class StorageVerificationError(StorageError):
    """Raised when the uploaded/master file cannot be verified."""


class GoogleAppsScriptStorage:
    """Read/write adapter that communicates with Google Apps Script Web App.

    The Apps Script Web App must implement the following endpoints:

    - POST /exec
      Body JSON: {"token": "<secreto>", "action": "get_master"}
      Returns: {"ok": true, "action": "get_master", "fileName": "BaseDatos_PQR.xlsx",
                "fileId": "...", "mimeType": "...", "size": N, "base64": "..."}

    - POST /exec
      Body JSON: {"token": "<secreto>", "action": "save_master",
                   "base64": "<contenido XLSX base64>"}
      Returns: {"ok": true, "action": "save_master"}

    - POST /exec
      Body JSON: {"token": "<secreto>", "action": "backup_master"}
      Returns: {"ok": true, "action": "backup_master"}

    - GET /exec?action=health
      Returns: {"ok": true, "status": "ok"}

    The Python module ONLY needs:
    - PQR_STORAGE_URL: URL base de la Apps Script (https://script.google.com/macros/s/EXEC_ID/exec)
    - PQR_STORAGE_TOKEN: Token secreto compartido entre Flask y la Apps Script
    - No Google credentials, no OAuth, no Service Account in the Python code
    """

    def __init__(self):
        self.base_url = os.getenv("PQR_STORAGE_URL", "").rstrip("/").strip()
        self.token = os.getenv("PQR_STORAGE_TOKEN", "").strip()

    def _get_session(self):
        """Crear y configurar una sesión HTTP robusta para Apps Script."""
        session = requests.Session()
        # Disable auto-redirects to preserve POST method.
        # Google Apps Script may redirect; we handle them manually to keep POST.
        session.allow_redirects = False
        # Reasonable timeout for Google Apps Script
        session.timeout = (5, 30)
        return session

    def _post_exec(self, action, data=None):
        """Llamada interna a los endpoints POST /exec de la Apps Script.

        El token se ENVÍA DENTRO del JSON body, NO en headers.
        Manejo robusto de respuestas HTTP y errores.
        Manejo manual de redirects para preservar el método POST.
        """
        if not self.base_url:
            raise StorageConfigurationError(
                "PQR_STORAGE_URL es obligatorio en modo google_drive."
            )
        if not self.token:
            raise StorageConfigurationError(
                "PQR_STORAGE_TOKEN es obligatorio en modo google_drive."
            )

        payload = {"token": self.token, "action": action}
        if data:
            payload.update(data)

        url = f"{self.base_url}"

        session = self._get_session()
        try:
            # Primera petición POST
            response = session.post(url, json=payload, timeout=30)
            http_status = response.status_code
            final_url = response.url

            # Capturar información de diagnóstico sin exponer tokens
            response_json = None
            response_error = None
            try:
                response_json = response.json()
                response_error = response_json.get("error")
            except Exception:
                response_json = None
                response_error = None

            # Determinar si es éxito (códigos 2xx)
            is_success = 200 <= http_status < 300

            # Manejo manual de redirects (301, 302, 303, 307, 308)
            # Google Apps Script puede redirigir; seguimos el Location preservando POST
            max_redirects = 5
            redirect_count = 0

            while not is_success and http_status in (301, 302, 303, 307, 308) and redirect_count < max_redirects:
                redirect_count += 1
                location = response.headers.get("Location", "")

                if not location:
                    break

                # Seguir redirect con POST (307/308 preservan método, 302/303 pueden cambiarlo)
                # Para Apps Script, realizamos otro POST al Location
                if http_status in (301, 302, 303):
                    # 302/303: Tradicionalmente cambian a GET, pero forzamos POST para Apps Script
                    response = session.post(location, json=payload, timeout=30)
                else:
                    # 307/308: Preservan el método POST
                    response = session.post(location, json=payload, timeout=30)

                http_status = response.status_code
                final_url = response.url

                # Intentar parsear respuesta JSON
                try:
                    response_json = response.json()
                    response_error = response_json.get("error") if response_json else None
                except Exception:
                    response_json = None
                    response_error = None

                is_success = 200 <= http_status < 300

            if not is_success:
                LOGGER.exception(
                    "Error HTTP %s en Google Apps Script durante: %s",
                    http_status, action
                )
                raise StorageUnavailableError(
                    f"Google Apps Script retornó error HTTP {http_status} durante: {action}."
                )

            # Si la respuesta no tiene ok: true, lanzar error con campo error
            if response_json and not response_json.get("ok", False):
                error_msg = response_json.get("error")
                if error_msg:
                    raise StorageVerificationError(
                        f"Google Apps Script error: {error_msg}"
                    )
                # Si no hay campo error específico, lanzar error genérico
                raise StorageUnavailableError(
                    f"Google Apps Script no pudo completar la operación: {action}. "
                    f"Respuesta: {response_json}"
                )

            return response_json

        except requests.exceptions.RequestException as error:
            LOGGER.exception("Error de conexión en Google Apps Script durante: %s", action)
            raise StorageUnavailableError(
                f"Error de conexión con Google Apps Script durante: {action}."
            ) from error
        except StorageConfigurationError:
            raise
        except StorageVerificationError:
            raise
        except StorageUnavailableError:
            raise
        except Exception as error:
            LOGGER.exception("Error inesperado en Google Apps Script durante: %s", action)
            raise StorageUnavailableError(
                f"Error inesperado en Google Apps Script durante: {action}."
            ) from error

    def _get_exec(self, action=None, params=None):
        """Llamada interna a los endpoints GET /exec de la Apps Script."""
        if not self.base_url:
            raise StorageConfigurationError(
                "PQR_STORAGE_URL es obligatorio en modo google_drive."
            )
        if not self.token:
            raise StorageConfigurationError(
                "PQR_STORAGE_TOKEN es obligatorio en modo google_drive."
            )

        url = self.base_url
        if action:
            url = f"{self.base_url}?action={action}"

        session = self._get_session()
        try:
            response = session.get(url, timeout=30)
            http_status = response.status_code
            final_url = response.url

            # Sanitizar respuesta
            response_json = None
            try:
                response_json = response.json()
            except Exception:
                response_json = None

            is_success = 200 <= http_status < 300

            if not is_success:
                LOGGER.exception(
                    "Error HTTP %s en Google Apps Script GET durante: %s",
                    http_status, action
                )
                raise StorageUnavailableError(
                    f"Google Apps Script retornó error HTTP {http_status} GET durante: {action}."
                )

            if response_json and not response_json.get("ok", False):
                error_msg = response_json.get("error")
                if error_msg:
                    raise StorageVerificationError(
                        f"Google Apps Script error GET: {error_msg}"
                    )
                raise StorageUnavailableError(
                    f"Google Apps Script no pudo completar la operación GET: {action}."
                )

            return response_json

        except requests.exceptions.RequestException as error:
            LOGGER.exception("Error de conexión en Google Apps Script GET durante: %s", action)
            raise StorageUnavailableError(
                f"Error de conexión con Google Apps Script GET durante: {action}."
            ) from error
        except StorageConfigurationError:
            raise
        except StorageVerificationError:
            raise
        except StorageUnavailableError:
            raise
        except Exception as error:
            LOGGER.exception("Error inesperado en Google Apps Script GET durante: %s", action)
            raise StorageUnavailableError(
                f"Error inesperado en Google Apps Script GET durante: {action}."
            ) from error

    def asegurar_maestro(self):
        """Confirm that the configured master exists without creating a replacement.

        Hace una petición GET health o get_master para validar que el archivo
        maestro está disponible en Google Drive.
        """
        # Primero intentamos health check
        try:
            health = self._get_exec("health")
            if health and health.get("ok"):
                return True
        except Exception:
            pass

        # Si no hay health, intentamos get_master
        try:
            master = self.get_master()
            return master.get("ok", False)
        except Exception:
            pass

        raise StorageUnavailableError(
            "No fue posible verificar la existencia del archivo maestro."
        )

    def get_master(self):
        """Descargar el Excel maestro desde Google Drive mediante Apps Script.

        Hace POST /exec con {"token": ..., "action": "get_master"}.
        El resultado incluye el contenido Base64 del archivo XLSX.

        Returns:
            dict: {"ok": true, "action": "get_master", "fileName": "...",
                   "fileId": "...", "mimeType": "...", "size": N, "base64": "..."}
        """
        result = self._post_exec("get_master")
        if not result.get("ok"):
            raise StorageVerificationError(
                "Google Apps Script no pudo obtener el archivo maestro."
            )
        return result

    def save_master(self, content_base64):
        """Subir/reemplazar el Excel maestro en Google Drive.

        Envía el contenido Base64 del archivo XLSX a la Apps Script,
        que lo guardará reemplazando el archivo maestro existente.

        Args:
            content_base64: Cadena Base64 del archivo XLSX a guardar.

        Returns:
            dict: {"ok": true, "action": "save_master"}
        """
        result = self._post_exec("save_master", {"base64": content_base64})
        if not result.get("ok"):
            raise StorageVerificationError(
                "Google Apps Script no confirmó el guardado del archivo maestro."
            )
        return result

    def backup_master(self):
        """Crear un backup del maestro en Google Drive.

        La Apps Script debe crear la copia en una carpeta de backups antes
        de reemplazar el archivo maestro.

        Returns:
            dict: {"ok": true, "action": "backup_master"}
        """
        result = self._post_exec("backup_master")
        if not result.get("ok"):
            raise StorageVerificationError(
                "Google Apps Script no pudo crear el backup del maestro."
            )
        return result

    def verificar_maestro(self, radicado=None, cantidad_esperada=None):
        """Validar el XLSX remoto después de subirlo.

        Descarga el archivo y valida:
        - Que pueda abrirse como XLSX válido (tiene estructura ZIP)
        - Que existan las hojas obligatorias: PQR, Historial, Investigaciones, Adjuntos, Usuarios
        - Que los encabezados de la hoja PQR coincidan exactamente
        - Que la cantidad de PQR sea la esperada
        - Que el radicado solicitado exista (opcional)

        Args:
            radicado: Radicado opcional a verificar que existe en el maestro
            cantidad_esperada: Cantidad esperada de PQR (opcional)

        Returns:
            dict: {"ok": true, "count": N, "radicado_exists": bool, "sha256": "..."}
        """
        # Descargar el maestro actual
        master = self.get_master()
        if not master.get("ok"):
            raise StorageVerificationError("No se pudo descargar el maestro para verificar.")

        content_base64 = master.get("base64", "")
        if not content_base64:
            raise StorageVerificationError("El maestro descargado no tiene contenido.")

        # Decodificar y validar el XLSX
        try:
            import base64 as b64mod
            import zipfile

            xlsx_bytes = b64mod.b64decode(content_base64)

            # Verificar que sea un XLSX válido (tiene estructura ZIP)
            if not zipfile.is_zipfile(xlsx_bytes):
                raise StorageVerificationError("El archivo maestro no es un XLSX válido.")

            # Leer con openpyxl verificando la estructura
            from openpyxl import load_workbook

            libro = load_workbook(xlsx_bytes, read_only=True, data_only=False)

            # Validar hojas obligatorias
            faltantes = [hoja for hoja in REQUIRED_SHEETS if hoja not in libro.sheetnames]
            if faltantes:
                libro.close()
                raise StorageVerificationError(
                    f"Faltan hojas obligatorias en el maestro: {', '.join(faltantes)}."
                )

            # Validar encabezados de la hoja PQR
            hoja_pqr = libro["PQR"]
            encabezados = [celda.value for celda in next(hoja_pqr.iter_rows(min_row=1, max_row=1))]
            if encabezados[:len(PQR_HEADERS)] != list(PQR_HEADERS):
                libro.close()
                raise StorageVerificationError("Los encabezados de la hoja PQR no coinciden.")

            # Contar radicados
            radicados = set()
            for fila in hoja_pqr.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
                valor = str(fila[0] or "").strip()
                if valor:
                    radicados.add(valor.upper())

            cantidad = len(radicados)
            existe = radicado is None or str(radicado).strip().upper() in radicados

            if cantidad_esperada is not None and cantidad != cantidad_esperada:
                libro.close()
                raise StorageVerificationError(
                    f"Se esperaban {cantidad_esperada} PQR y se encontraron {cantidad}."
                )
            if not existe and radicado is not None:
                libro.close()
                raise StorageVerificationError(
                    f"El radicado {radicado} no existe en el maestro verificado."
                )

            resultado = {
                "ok": True,
                "count": cantidad,
                "radicado_exists": existe,
                "sha256": hashlib.sha256(xlsx_bytes).hexdigest(),
                "sheets": list(libro.sheetnames)
            }
            libro.close()
            return resultado

        except Exception as error:
            LOGGER.exception("Error al validar el maestro verificado")
            raise StorageVerificationError(
                "No fue posible validar el archivo maestro."
            ) from error

    def subir_evidencia(self, origen, nombre=None):
        """NOTA: La API actual de Google Apps Script NO tiene el endpoint save_evidence.

        Este método está dejado como placeholder/pausado. Subir evidencia
        requiere implementar un nuevo endpoint en la Apps Script Web App
        que no forma parte del scope actual.

        Args:
            origen: Ruta local al archivo a subir (no usado actualmente)
            nombre: Nombre opcional (no usado actualmente)

        Returns:
            dict: {"ok": false, "mensaje": "Endpoint no implementado aún"}
        """
        return {"ok": False, "mensaje": "Endpoint save_evidence no implementado. Verificar API Apps Script."}


# Función de ayuda para usar sin crear instancia cada vez
_storage_adapter = None


def obtener_storage():
    """Obtener el adaptador de almacenamiento Google Apps Script."""
    global _storage_adapter

    modo = os.getenv("PQR_STORAGE", "local").strip().lower() or "local"
    if modo == "local":
        return None
    if modo != "google_drive":
        raise StorageConfigurationError(
            f"PQR_STORAGE no soportado: {modo}. Use local o google_drive."
        )

    if _storage_adapter is None:
        _storage_adapter = GoogleAppsScriptStorage()
    return _storage_adapter
