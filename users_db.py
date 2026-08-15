import os
from datetime import datetime
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash, check_password_hash

from excel_db import (
    ARCHIVO,
    cargar_workbook_seguro,
    crear_excel,
    guardar_workbook_atomico,
    proteger_escritura
)

print(">>> USANDO users_db.py <<<")

# ==========================================================
# ROLES
# ==========================================================

ADMIN = "ADMIN"
VENDEDOR = "VENDEDOR"
LIDER_CALIDAD = "LIDER_CALIDAD"
LIDER_COMERCIAL = "LIDER_COMERCIAL"
COORDINADORA_COMERCIAL = "COORDINADORA COMERCIAL"
DIRECTORA_COMERCIAL = "DIRECTORA COMERCIAL"
COMERCIAL = "COMERCIAL"
DIRECTOR_PRODUCCION = "DIRECTOR DE PRODUCCION"

ROLES_VALIDOS = (
    ADMIN,
    VENDEDOR,
    LIDER_CALIDAD,
    LIDER_COMERCIAL,
    COORDINADORA_COMERCIAL,
    DIRECTORA_COMERCIAL,
    COMERCIAL,
    DIRECTOR_PRODUCCION
)

# ==========================================================
# HOJA USUARIOS
# ==========================================================

COLUMNAS = [
    "ID",
    "Nombre",
    "Usuario",
    "ContrasenaHash",
    "Rol",
    "LineaProducto",
    "Empresa",
    "Activo",
    "FechaCreacion",
    "Documento",
    "Correo",
    "Telefono"
]


@proteger_escritura
def _asegurar_hoja():

    if not os.path.exists(ARCHIVO):
        crear_excel()

    wb = cargar_workbook_seguro()
    cambios = False

    if "Usuarios" not in wb.sheetnames:
        ws = wb.create_sheet("Usuarios")
        ws.append(COLUMNAS)
        cambios = True
    else:
        ws = wb["Usuarios"]
        if ws.max_column < 10:
            ws.cell(1, 10).value = "Documento"
            cambios = True
        if ws.max_column < 11:
            ws.cell(1, 11).value = "Correo"
            cambios = True
        if ws.max_column < 12:
            ws.cell(1, 12).value = "Telefono"
            cambios = True

    if cambios:
        guardar_workbook_atomico(wb)
    else:
        wb.close()


def _fila_a_usuario(fila):

    activo = True
    if len(fila) > 7 and fila[7] is not None:
        activo = str(fila[7]).strip() in ("1", "True", "true", "Sí", "SI", "Activo")

    return {
        "id": fila[0],
        "nombre": fila[1] or "",
        "usuario": fila[2] or "",
        "documento": fila[9] if len(fila) > 9 and fila[9] else "",
        "rol": fila[4] or "",
        "linea_producto": fila[5] or "",
        "empresa": fila[6] or "INAPEL",
        "activo": activo,
        "fecha_creacion": str(fila[8]) if len(fila) > 8 and fila[8] else "",
        "correo": fila[10] if len(fila) > 10 and fila[10] else "",
        "telefono": fila[11] if len(fila) > 11 and fila[11] else ""
    }


def _existe_usuario(usuario):

    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    existe = False

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[2] and str(fila[2]).strip().lower() == str(usuario).strip().lower():
            existe = True
            break

    wb.close()
    return existe


def _existe_documento(documento, excepto_uid=None):

    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    existe = False

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[9] and str(fila[9]).strip() == str(documento).strip():
            if excepto_uid is not None and str(fila[0]).strip() == str(excepto_uid).strip():
                continue
            existe = True
            break

    wb.close()
    return existe


def _existe_correo(correo, excepto_uid=None):

    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    existe = False

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[10] and str(fila[10]).strip().lower() == str(correo).strip().lower():
            if excepto_uid is not None and str(fila[0]).strip() == str(excepto_uid).strip():
                continue
            existe = True
            break

    wb.close()
    return existe


def documento_disponible(documento, excepto_uid=None):
    return not _existe_documento(documento, excepto_uid)


def correo_disponible(correo, excepto_uid=None):
    return not _existe_correo(correo, excepto_uid)


def _normalizar_telefono(telefono):
    return "".join(ch for ch in str(telefono or "") if ch.isdigit())


def _siguiente_id():

    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    max_id = 0

    for fila in ws.iter_rows(min_row=2, values_only=True):
        try:
            max_id = max(max_id, int(fila[0]))
        except Exception:
            pass

    wb.close()
    return max_id + 1


# ==========================================================
# CREAR USUARIO
# ==========================================================

@proteger_escritura
def crear_usuario(
    nombre,
    usuario,
    contrasena,
    rol,
    documento="",
    linea_producto="",
    empresa="INAPEL",
    activo=True,
    correo="",
    telefono=""
):

    if _existe_usuario(usuario):
        return {"ok": False, "mensaje": "El usuario ya existe."}

    if documento and _existe_documento(documento):
        return {"ok": False, "mensaje": "El documento ya está registrado."}

    if correo and _existe_correo(correo):
        return {"ok": False, "mensaje": "El correo electrónico ya está registrado."}

    if rol not in ROLES_VALIDOS:
        return {"ok": False, "mensaje": "Rol inválido."}

    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    nuevo_id = _siguiente_id()

    ws.append([
        nuevo_id,
        nombre,
        usuario,
        generate_password_hash(contrasena),
        rol,
        linea_producto,
        empresa,
        1 if activo else 0,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        documento,
        correo,
        _normalizar_telefono(telefono)
    ])

    guardar_workbook_atomico(wb)

    return {"ok": True, "id": nuevo_id, "mensaje": "Usuario creado correctamente."}


# ==========================================================
# AUTENTICAR
# ==========================================================

def autenticar_usuario(usuario, contrasena):

    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    for fila in ws.iter_rows(min_row=2, values_only=True):

        if not fila[2]:
            continue

        if str(fila[2]).strip().lower() != str(usuario).strip().lower():
            continue

        activo = True
        if fila[7] is not None:
            activo = str(fila[7]).strip() in ("1", "True", "true", "Sí", "SI", "Activo")

        if not activo:
            wb.close()
            return {"error": "Usuario inactivo. Contacte al administrador."}

        if not fila[3] or not check_password_hash(fila[3], contrasena):
            wb.close()
            return {"error": "Credenciales incorrectas."}

        usuario_data = _fila_a_usuario(fila)
        wb.close()
        return usuario_data

    wb.close()
    return {"error": "Credenciales incorrectas."}


# ==========================================================
# LISTAR
# ==========================================================

def listar_usuarios():

    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    lista = []

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila[2]:
            continue
        lista.append(_fila_a_usuario(fila))

    wb.close()
    return lista


def obtener_usuario_por_id(uid):

    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0] is not None and str(fila[0]).strip() == str(uid).strip():
            wb.close()
            return _fila_a_usuario(fila)

    wb.close()
    return None


def obtener_usuario_por_documento(documento):

    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[9] is not None and str(fila[9]).strip() == str(documento).strip():
            wb.close()
            return _fila_a_usuario(fila)

    wb.close()
    return None


def usuario_disponible(usuario, excepto_uid=None):
    """
    Verifica que el nombre de usuario no esté en uso por otro usuario.
    Si excepto_uid se indica, ese usuario se ignora en la comprobación
    (permite conservar el propio usuario al editar credenciales).
    """

    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    for fila in ws.iter_rows(min_row=2, values_only=True):
        if not fila[2]:
            continue
        if str(fila[2]).strip().lower() != str(usuario).strip().lower():
            continue
        if excepto_uid is not None and str(fila[0]).strip() == str(excepto_uid).strip():
            continue
        wb.close()
        return False

    wb.close()
    return True


# ==========================================================
# ACTUALIZAR
# ==========================================================

@proteger_escritura
def actualizar_usuario(uid, **campos):

    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    encontrado = False

    for fila in ws.iter_rows(min_row=2):

        if str(fila[0].value).strip() != str(uid).strip():
            continue

        encontrado = True

        if "nombre" in campos:
            fila[1].value = campos["nombre"]
        if "usuario" in campos:
            fila[2].value = campos["usuario"]
        if "contrasena" in campos and campos["contrasena"]:
            fila[3].value = generate_password_hash(campos["contrasena"])
        if "rol" in campos:
            fila[4].value = campos["rol"]
        if "linea_producto" in campos:
            fila[5].value = campos["linea_producto"]
        if "empresa" in campos:
            fila[6].value = campos["empresa"]
        if "activo" in campos:
            fila[7].value = 1 if campos["activo"] else 0
        if "documento" in campos:
            fila[9].value = campos["documento"]
        if "correo" in campos:
            fila[10].value = campos["correo"]
        if "telefono" in campos:
            fila[11].value = _normalizar_telefono(campos["telefono"])

        break

    if not encontrado:
        wb.close()
        return {"ok": False, "mensaje": "Usuario no encontrado."}

    guardar_workbook_atomico(wb)

    return {"ok": True, "mensaje": "Usuario actualizado correctamente."}


def desactivar_usuario(uid):
    return actualizar_usuario(uid, activo=False)


@proteger_escritura
def eliminar_usuario(uid):
    """
    Elimina físicamente el registro del usuario de la hoja Usuarios.
    No toca PQR, Historial, Investigaciones ni Adjuntos.
    """
    _asegurar_hoja()

    wb = cargar_workbook_seguro()
    ws = wb["Usuarios"]

    numero_fila = None

    for fila in ws.iter_rows(min_row=2):
        if fila[0].value is not None and str(fila[0].value).strip() == str(uid).strip():
            numero_fila = fila[0].row
            break

    if numero_fila is None:
        wb.close()
        return {"ok": False, "mensaje": "Usuario no encontrado."}

    ws.delete_rows(numero_fila, 1)

    guardar_workbook_atomico(wb)

    return {"ok": True, "mensaje": "Usuario eliminado correctamente."}


# ==========================================================
# SEMBRAR USUARIOS AL INICIAR LA APLICACIÓN
# ==========================================================

def sembrar_usuarios():
    """
    1) Crea el administrador general (preserva el acceso actual del sistema).
       La contraseña se obtiene exclusivamente de la variable de entorno
       ADMIN_PASS.
    2) Crea los usuarios definidos en usuarios_iniciales.py (estructura clara
       para registrar los 14 vendedores y los 3 líderes cuando se dispongan
       los nombres y credenciales definitivos).
    """

    if not _existe_usuario("admin"):
        clave = os.getenv("ADMIN_PASS")
        if not clave:
            raise RuntimeError(
                "ADMIN_PASS debe configurarse antes de crear el administrador inicial."
            )
        crear_usuario(
            nombre="Administrador General",
            usuario="admin",
            contrasena=clave,
            rol=ADMIN,
            linea_producto="",
            empresa="INAPEL",
            activo=True
        )

    try:
        from usuarios_iniciales import USUARIOS_INICIALES
    except ImportError:
        USUARIOS_INICIALES = []

    for u in USUARIOS_INICIALES:

        usuario_login = str(u.get("usuario", "")).strip()

        if not usuario_login:
            continue

        if _existe_usuario(usuario_login):
            continue

        crear_usuario(
            nombre=str(u.get("nombre", "")).strip(),
            usuario=usuario_login,
            contrasena=str(u.get("contrasena", "")),
            rol=str(u.get("rol", "")).strip().upper(),
            documento=str(u.get("documento", "")).strip(),
            linea_producto=str(u.get("linea_producto", "")).strip().upper(),
            empresa=str(u.get("empresa", "") or "INAPEL").strip().upper(),
            activo=u.get("activo", True)
        )
